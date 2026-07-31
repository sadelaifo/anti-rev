#!/usr/bin/env python3
"""
pyarmor_compat_scan  -  statically flag scripts that PyArmor may break.

PyArmor obfuscates by transforming/encrypting bytecode and installing a
runtime import hook.  Anything that reads a function's *real* bytecode or
source, ships a *function* to another process, or (in rename mode) resolves
names dynamically can break.  This tool parses every ``.py`` under a tree with
the ``ast`` module and reports the at-risk patterns, ranked by severity, so you
know which modules to EXCLUDE from obfuscation (or feed to
``encryptor/pyarmor_verify.py`` for a behavioural check).

It is a PRE-FILTER, not a proof of safety:
  * it catches syntactic patterns (import X, @decorator, call foo())
  * it CANNOT catch dynamic cases (getattr(m, name_from_config)(), a function
    pickled only under a runtime branch, spawn chosen via env var).
Treat the output as a triage list; confirm behaviour with pyarmor_verify.py.

eval/exec provenance: exec/eval/compile findings are traced back (best-effort,
within one file) to where their input comes from. Input that traces to a
literal or a local/config file is OMITTED (trusted, and PyArmor runs it fine);
input from network/user/env (requests, input(), response.text, ...) or of
undetermined origin is kept as a code-injection (RCE) risk. --no-trace-exec
restores the old "flag everything" behaviour.

exec/eval of a CONSTRUCTED code string (f-string/concat/.format that builds
source from live variable names, e.g. exec(f'{foo}{field} = v')) is reported
separately as 'dynamic-codegen': it runs fine under basic obfuscation but
BREAKS under RFT rename, since the names in the string aren't renamed with the
variables. These are always kept (never omitted) for review.

Scope (--checks, default 'pyarmor'):
  By default this reports ONLY failures that OBFUSCATION INTRODUCES - code that
  works in plaintext but breaks once obfuscated (numba, inspect.getsource, dill/
  cloudpickle, open(__file__) self-read, loader resources, reload). Pre-existing
  business-logic / security issues that fail the SAME with or without PyArmor -
  spawn/pool picklability, eval-of-untrusted-input (RCE), sibling data-file
  loads - are the developer's responsibility and are HIDDEN. Pass --checks all
  to include them as an extra lint.

Severity:
  HIGH    almost certainly breaks under obfuscation -> exclude the module
  MEDIUM  breaks in common configs -> review
  LOW     only matters in specific modes (RFT rename / BCC) -> mode-dependent
  INFO    worth knowing, usually fine

Usage:
  tools/pyarmor_compat_scan.py /path/to/pysuite
  tools/pyarmor_compat_scan.py . --json > report.json
  tools/pyarmor_compat_scan.py src --xlsx report.xlsx    # Excel workbook
  tools/pyarmor_compat_scan.py src --min-severity MEDIUM
  tools/pyarmor_compat_scan.py src --no-rft --no-bcc      # basic mode only
  tools/pyarmor_compat_scan.py src --exclude tests --exclude build
  tools/pyarmor_compat_scan.py huge/tree -j 8             # 8 parallel workers

Exit code is non-zero when a finding at or above --fail-on (default HIGH) is
present, so it can gate CI.
"""
from __future__ import annotations

import argparse
import ast
import json
import os
import sys
from pathlib import Path

SEVERITY_ORDER = {"INFO": 0, "LOW": 1, "MEDIUM": 2, "HIGH": 3}

# Categories that are NOT PyArmor-introduced: the code behaves the SAME with or
# without obfuscation (pre-existing business logic / security). Shown only with
# --checks all; hidden by the default --checks pyarmor.
BUSINESS_CATEGORIES = {"dynamic-exec", "func-dispatch", "resource-sibling"}

# --- dotted callables that are hard breakers, keyed to a category ---
# JIT / bytecode-compiling decorators: PyArmor-transformed bytecode is
# unreadable to these compilers (see the Numba "analyzing bytecode" failure).
JIT_CALLABLES = {
    "numba.jit", "numba.njit", "numba.vectorize", "numba.guvectorize",
    "numba.stencil", "numba.cfunc", "numba.cuda.jit", "numba.experimental.jitclass",
}
# Reading source / bytecode of live objects -> nothing there after obfuscation.
SOURCE_CALLABLES = {
    "inspect.getsource", "inspect.getsourcelines", "inspect.getsourcefile",
    "inspect.getcomments", "dis.dis", "dis.Bytecode", "dis.get_instructions",
}
# marshal round-trips code objects; blocked / meaningless once obfuscated.
MARSHAL_CALLABLES = {"marshal.dumps", "marshal.loads"}
# Anti-debug commonly disables tracing -> profilers/coverage/debuggers die.
TRACE_CALLABLES = {"sys.settrace", "sys.setprofile", "threading.settrace", "threading.setprofile"}
# Ships a FUNCTION across a process boundary -> can't marshal obfuscated code.
FUNC_SERIALIZE_CALLABLES = {
    "multiprocessing.set_start_method", "multiprocessing.get_context",
    "concurrent.futures.ProcessPoolExecutor",
}
# Dynamic code paths  -  usually fine, but interact with import hooks / BCC.
DYNAMIC_EXEC_NAMES = {"exec", "eval", "compile"}
RFT_ATTR_BUILTINS = {"getattr", "setattr", "hasattr", "delattr"}

# --- #4 resource-path / loader assumptions ---
# Obfuscation changes __file__/__loader__; packaged-data reads via the loader
# can fail, and paths computed from __file__ may point at the obfuscated file.
RESOURCE_CALLABLES = {
    "pkg_resources.resource_filename", "pkg_resources.resource_string",
    "pkg_resources.resource_stream", "pkg_resources.resource_listdir",
    "importlib.resources.files", "importlib.resources.path",
    "importlib.resources.read_text", "importlib.resources.read_binary",
    "importlib.resources.open_text", "importlib.resources.open_binary",
    "pkgutil.get_data",
}
# path builders that are a problem only when fed __file__
PATH_FROM_FILE_CALLABLES = {
    "os.path.dirname", "os.path.abspath", "os.path.realpath", "os.path.join",
    "pathlib.Path",
}
FRAME_INTROSPECT_CALLABLES = {
    "inspect.stack", "inspect.currentframe", "inspect.getframeinfo",
    "inspect.trace",
}
# --- #6 dynamic name resolution (RFT rename mode) ---
RFT_NAME_CALLABLES = {"operator.attrgetter", "operator.methodcaller"}

# --- eval/exec source provenance (trace where the executed string comes from) ---
# EXTERNAL = untrusted / attacker-influenceable -> keep (code-injection risk).
EXTERNAL_SOURCE_CALLS = {
    "input", "requests.get", "requests.post", "requests.put", "requests.patch",
    "requests.delete", "requests.head", "requests.request",
    "urllib.request.urlopen", "httpx.get", "httpx.post", "httpx.request",
    "os.getenv", "os.environ.get",
    "subprocess.check_output", "subprocess.getoutput", "subprocess.run",
}
EXTERNAL_METHODS = {"recv", "recvfrom"}                  # socket reads
EXTERNAL_ATTRS = {"text", "content", "body", "raw"}      # response.text/.content/...
# LOCAL = literal or in-tree/config file -> omit (trusted, shipped alongside code).
LOCAL_SOURCE_CALLS = {
    "open", "pathlib.Path", "json.load", "json.loads",
    "yaml.load", "yaml.safe_load", "toml.load",
}
LOCAL_METHODS = {"read", "read_text", "readlines", "getvalue"}

# module imports that are, on their own, a signal
IMPORT_SIGNALS = {
    # module -> (severity, category, note)
    "numba": ("HIGH", "numba-jit", "Numba JIT  -  obfuscated bytecode is unreadable to it; exclude jitted funcs"),
    "coverage": ("HIGH", "profiler", "coverage.py uses settrace; PyArmor anti-debug blocks it"),
    "line_profiler": ("HIGH", "profiler", "line_profiler uses settrace; blocked by anti-debug"),
    "memory_profiler": ("MEDIUM", "profiler", "memory_profiler traces; may be blocked"),
    "dill": ("HIGH", "func-serialization", "dill pickles functions/code  -  fails on obfuscated code objects"),
    "cloudpickle": ("HIGH", "func-serialization", "cloudpickle exists to serialize functions  -  fails on obfuscated code"),
    "celery": ("MEDIUM", "func-dispatch", "Celery dispatches tasks by NAME; the worker imports the (obfuscated) module  -  works either way, not PyArmor-introduced"),
    "joblib": ("MEDIUM", "func-dispatch", "joblib may pickle a callable to a worker  -  a top-level func works, a lambda fails with or without PyArmor"),
    "Cython": ("INFO", "compiled", "Cython output is already compiled  -  can't/needn't be obfuscated"),
    "cython": ("INFO", "compiled", "Cython output is already compiled  -  can't/needn't be obfuscated"),
    "sqlalchemy": ("LOW", "rft", "ORM maps columns to attribute names  -  RFT rename can break mapping"),
    "django": ("LOW", "rft", "Django ORM/forms map fields to attribute names  -  RFT rename risk"),
    "pytest": ("MEDIUM", "test", "pytest rewrites test bytecode for asserts  -  don't obfuscate tests"),
    "sphinx": ("MEDIUM", "source-introspection", "Sphinx autodoc reads source/docstrings  -  gone after obfuscation"),
}


def dotted_name(node):
    """Return the dotted attribute chain for a Name/Attribute node, else None."""
    parts = []
    while isinstance(node, ast.Attribute):
        parts.append(node.attr)
        node = node.value
    if isinstance(node, ast.Name):
        parts.append(node.id)
        return ".".join(reversed(parts))
    return None


class Scanner(ast.NodeVisitor):
    def __init__(self, path, opts, assignments=None):
        self.path = path
        self.opts = opts
        self.assignments = assignments or {}   # var name -> [RHS AST nodes] (whole file)
        self.trace_exec = not getattr(opts, "no_trace_exec", False)
        self.checks_all = getattr(opts, "checks", "pyarmor") == "all"
        self.aliases = {}          # local name -> full dotted module/callable path
        self.findings = []         # (line, severity, category, message)
        self._imported_mods = set()

    # ---- finding helper -------------------------------------------------
    def add(self, line, severity, category, message):
        # RFT-only breakage (incl. constructed-code exec) is only introduced when
        # rename mode is on; BCC-only likewise. Suppressed when those are off.
        if category in ("rft", "dynamic-codegen") and self.opts.no_rft:
            return
        if category == "bcc" and self.opts.no_bcc:
            return
        # Not-PyArmor-introduced (business/security): only with --checks all.
        if category in BUSINESS_CATEGORIES and not self.checks_all:
            return
        self.findings.append((line, severity, category, message))

    # ---- resolve a call/decorator node to its full dotted path ----------
    def resolve(self, node):
        if isinstance(node, ast.Call):
            node = node.func
        dn = dotted_name(node)
        if dn is None:
            return None
        head, _, rest = dn.partition(".")
        if head in self.aliases:
            base = self.aliases[head]
            return base if not rest else f"{base}.{rest}"
        return dn

    # ---- imports: build the alias map ----------------------------------
    def visit_Import(self, node):
        for a in node.names:
            local = a.asname or a.name.split(".")[0]
            full = a.name if a.asname else a.name.split(".")[0]
            self.aliases[local] = a.name if a.asname else full
            top = a.name.split(".")[0]
            self._imported_mods.add(top)
            self._flag_import(top, node.lineno)
        self.generic_visit(node)

    def visit_ImportFrom(self, node):
        mod = node.module or ""
        top = mod.split(".")[0]
        if mod:
            self._imported_mods.add(top)
            self._flag_import(top, node.lineno)
        for a in node.names:
            local = a.asname or a.name
            self.aliases[local] = f"{mod}.{a.name}" if mod else a.name
        # antirev_client's own import hook can race PyArmor's import hook
        if top == "antirev_client" or mod == "antirev_client":
            self.add(node.lineno, "MEDIUM", "import-hook",
                     "imports antirev_client  -  its sys.meta_path/ctypes patch may "
                     "interact with PyArmor's runtime import hook; test this path")
        self.generic_visit(node)

    def _flag_import(self, top, line):
        if top in IMPORT_SIGNALS:
            sev, cat, note = IMPORT_SIGNALS[top]
            self.add(line, sev, cat, f"imports '{top}': {note}")

    # ---- decorators (the actionable "exclude this function" signal) -----
    def _check_decorators(self, node):
        for dec in node.decorator_list:
            full = self.resolve(dec)
            if full in JIT_CALLABLES:
                self.add(dec.lineno, "HIGH", "numba-jit",
                         f"@{full.split('.')[-1]} on '{node.name}()'  -  JIT compiler "
                         f"reads real bytecode; EXCLUDE this function from obfuscation")
            elif full in ("celery.shared_task",) or (full and full.endswith(".task")):
                self.add(dec.lineno, "MEDIUM", "func-dispatch",
                         f"task decorator on '{node.name}()'  -  Celery dispatches by name and "
                         f"the worker imports the obfuscated module; works, not PyArmor-introduced")

    def visit_FunctionDef(self, node):
        self._check_decorators(node)
        self._check_bcc(node)
        self.generic_visit(node)

    def visit_AsyncFunctionDef(self, node):
        self._check_decorators(node)
        if not self.opts.no_bcc:
            self.add(node.lineno, "LOW", "bcc",
                     f"'async def {node.name}'  -  some PyArmor BCC/bytecode-to-C versions "
                     f"don't support async; review if using BCC mode")
        self.generic_visit(node)

    def _check_bcc(self, node):
        # BCC (bytecode-to-C) mode has feature gaps; flag constructs to review.
        # Skip the (expensive) subtree walk entirely when BCC checks are off.
        if self.opts.no_bcc:
            return
        seen_yield = seen_nonlocal = False
        for n in ast.walk(node):
            if not seen_yield and isinstance(n, (ast.Yield, ast.YieldFrom)):
                self.add(n.lineno, "LOW", "bcc",
                         f"generator (yield) in '{node.name}'  -  verify under BCC mode")
                seen_yield = True
            elif not seen_nonlocal and isinstance(n, ast.Nonlocal):
                self.add(n.lineno, "LOW", "bcc",
                         f"nonlocal/closure in '{node.name}'  -  verify under BCC mode")
                seen_nonlocal = True
            if seen_yield and seen_nonlocal:
                break

    # ---- calls ----------------------------------------------------------
    def visit_Call(self, node):
        full = self.resolve(node)
        if full in JIT_CALLABLES:
            self.add(node.lineno, "HIGH", "numba-jit",
                     f"{full}() call  -  JIT reads real bytecode; exclude the target")
        elif full in SOURCE_CALLABLES:
            self.add(node.lineno, "HIGH", "source-introspection",
                     f"{full}()  -  reads source/bytecode that obfuscation removes")
        elif full in MARSHAL_CALLABLES:
            self.add(node.lineno, "MEDIUM", "bytecode-introspection",
                     f"{full}()  -  marshals code objects; meaningless/blocked once obfuscated")
        elif full in TRACE_CALLABLES:
            self.add(node.lineno, "HIGH", "trace-hooks",
                     f"{full}()  -  PyArmor anti-debug commonly disables tracing")
        elif full in FUNC_SERIALIZE_CALLABLES:
            self._check_spawn(node, full)
        elif full == "inspect.signature" or full == "inspect.getfullargspec":
            self.add(node.lineno, "LOW", "source-introspection",
                     f"{full}()  -  usually OK, but breaks if RFT renames params/args")
        elif full in ("importlib.reload",):
            self.add(node.lineno, "LOW", "import-hook",
                     "importlib.reload() on an obfuscated module can misbehave")
        elif full in RESOURCE_CALLABLES:
            self.add(node.lineno, "MEDIUM", "resource-path",
                     f"{full}()  -  reads packaged data via the loader; obfuscation "
                     f"changes __loader__/__file__ and this can fail")
        elif full in FRAME_INTROSPECT_CALLABLES:
            self.add(node.lineno, "LOW", "source-introspection",
                     f"{full}()  -  frame/source introspection; line/source info is "
                     f"degraded after obfuscation")
        elif full in RFT_NAME_CALLABLES:
            self.add(node.lineno, "LOW", "rft",
                     f"{full}()  -  resolves attributes by name; RFT rename mode may break it")
        elif full in PATH_FROM_FILE_CALLABLES and self._refs_dunder_file(node):
            self.add(node.lineno, "LOW", "resource-sibling",
                     f"{full}(__file__ ...)  -  locates a file next to the module; resolves "
                     f"the same obfuscated as plaintext if the data file ships alongside")

        # builtins (not import-resolved)
        fn = node.func
        if isinstance(fn, ast.Name):
            if fn.id in DYNAMIC_EXEC_NAMES:
                if not self.trace_exec:
                    self.add(node.lineno, "MEDIUM", "dynamic-exec",
                             f"{fn.id}()  -  dynamic code; interacts with PyArmor import hook / BCC")
                else:
                    arg = node.args[0] if node.args else None
                    src = self._classify_source(arg, frozenset())
                    if src == "external":                    # RCE wins (highest concern)
                        self.add(node.lineno, "MEDIUM", "dynamic-exec",
                                 f"{fn.id}() on EXTERNAL input (network/user/env)  -  "
                                 f"code-injection (RCE) risk; use json.loads / ast.literal_eval")
                    elif self._is_constructed_code(arg):
                        self.add(node.lineno, "MEDIUM", "dynamic-codegen",
                                 f"{fn.id}() on a CONSTRUCTED code string (f-string/concat/"
                                 f".format built from live variable names)  -  runs fine under "
                                 f"basic obfuscation but BREAKS under RFT rename; review")
                    elif src == "local":
                        pass   # input traces to a literal / config / local file -> omit
                    else:
                        self.add(node.lineno, "MEDIUM", "dynamic-exec",
                                 f"{fn.id}() input origin undetermined  -  verify it's trusted "
                                 f"(not network/user input); use ast.literal_eval if it's data")
            elif fn.id in RFT_ATTR_BUILTINS and len(node.args) >= 2:
                name_arg = node.args[1]
                is_literal = isinstance(name_arg, ast.Constant) and isinstance(name_arg.value, str)
                if not is_literal:
                    self.add(node.lineno, "LOW", "rft",
                             f"{fn.id}() with a computed name  -  RFT rename mode may break "
                             f"dynamic attribute access")
            elif fn.id in ("globals", "vars", "locals"):
                self.add(node.lineno, "LOW", "rft",
                         f"{fn.id}()  -  name-keyed access; RFT rename mode may break lookups")
            elif fn.id == "open" and self._refs_dunder_file(node):
                if self._targets_sibling_of_file(node):
                    self.add(node.lineno, "LOW", "resource-sibling",
                             "open() of a file next to the module (dirname(__file__)/..)  -  "
                             "resolves the same obfuscated as plaintext if the data file ships")
                else:
                    self.add(node.lineno, "MEDIUM", "resource-path",
                             "open(__file__)  -  reads the module's OWN source; after "
                             "obfuscation that path holds the obfuscated bytes, not your code")
            elif fn.id == "type" and len(node.args) == 3:
                self.add(node.lineno, "LOW", "rft",
                         "type(name, bases, dict)  -  dynamic class creation is name-based; "
                         "RFT rename mode may affect name-keyed registries")
        elif isinstance(fn, ast.Attribute) and fn.attr == "get_data":
            self.add(node.lineno, "LOW", "resource-path",
                     ".get_data()  -  loader data access (e.g. __loader__.get_data); "
                     "obfuscation changes the loader")
        self.generic_visit(node)

    def _check_spawn(self, node, full):
        if full == "concurrent.futures.ProcessPoolExecutor":
            self.add(node.lineno, "MEDIUM", "func-dispatch",
                     "ProcessPoolExecutor pickles the target by NAME  -  a top-level func "
                     "works (child re-imports the obfuscated module); a lambda/local fails "
                     "with or without PyArmor")
            return
        # multiprocessing.set_start_method('spawn') / get_context('spawn')
        for a in list(node.args) + [k.value for k in node.keywords]:
            if isinstance(a, ast.Constant) and a.value == "spawn":
                self.add(node.lineno, "MEDIUM", "func-dispatch",
                         "multiprocessing 'spawn' pickles the target by NAME  -  a top-level "
                         "func works, a lambda/local fails with or without PyArmor")
                return

    @staticmethod
    def _refs_dunder_file(node):
        return any(isinstance(n, ast.Name) and n.id == "__file__"
                   for n in ast.walk(node))

    @staticmethod
    def _targets_sibling_of_file(node):
        """True if the path goes through dirname(__file__)/Path(__file__).parent,
        i.e. it targets a SIBLING file rather than reading __file__ itself."""
        for n in ast.walk(node):
            if isinstance(n, ast.Attribute) and n.attr == "parent":
                return True
            if isinstance(n, ast.Call):
                f = n.func
                if isinstance(f, ast.Attribute) and f.attr in ("dirname", "split", "join"):
                    return True
                if isinstance(f, ast.Name) and f.id in ("dirname", "join"):
                    return True
        return False

    @staticmethod
    def _is_constructed_code(node):
        """True if the exec/eval arg is a code string BUILT at runtime (f-string,
        concat, %-format, .format()/.join()) rather than a plain literal/var."""
        if isinstance(node, ast.JoinedStr):                 # f'{foo} = ...'
            return any(isinstance(v, ast.FormattedValue) for v in node.values)
        if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Mod)):
            return True                                      # 'x' + name  /  '%s' % name
        if (isinstance(node, ast.Call) and isinstance(node.func, ast.Attribute)
                and node.func.attr in ("format", "join")):
            return True                                      # '{}'.format(name) / ''.join(...)
        return False

    # ---- eval/exec source provenance (best-effort, intra-file) ----------
    def _classify_source(self, node, seen):
        """Where does a value come from? 'local' (trusted) | 'external' | 'unknown'."""
        if node is None:
            return "unknown"
        if isinstance(node, ast.Constant):
            return "local"                       # a literal is in-source, trusted
        if isinstance(node, ast.Name):
            return self._classify_name(node.id, seen)
        if isinstance(node, ast.Call):
            return self._classify_call(node, seen)
        if isinstance(node, ast.Attribute):
            return "external" if node.attr in EXTERNAL_ATTRS else "unknown"
        if isinstance(node, ast.JoinedStr):      # f-string
            return self._combine(self._classify_source(v.value, seen)
                                 for v in node.values
                                 if isinstance(v, ast.FormattedValue))
        if isinstance(node, ast.BinOp):          # concat / % formatting
            return self._combine([self._classify_source(node.left, seen),
                                  self._classify_source(node.right, seen)])
        return "unknown"

    @staticmethod
    def _combine(classes):
        classes = set(classes)
        if "external" in classes:                # any tainted part taints the whole
            return "external"
        if classes and classes <= {"local"}:
            return "local"
        return "unknown"

    def _classify_name(self, name, seen):
        if name in seen:                         # guard assignment cycles (a=b; b=a)
            return "unknown"
        rhss = self.assignments.get(name)
        if not rhss:                             # param / import / builtin / not seen
            return "unknown"
        seen = seen | {name}
        return self._combine(self._classify_source(r, seen) for r in rhss)

    def _classify_call(self, node, seen):
        full = self.resolve(node)
        if full in EXTERNAL_SOURCE_CALLS:
            return "external"
        if full in LOCAL_SOURCE_CALLS:
            return "local"
        if isinstance(node.func, ast.Attribute):
            meth = node.func.attr
            if meth in EXTERNAL_METHODS or meth == "json":   # sock.recv(), resp.json()
                return "external"
            if meth in LOCAL_METHODS:            # f.read()/.read_text(): trust = receiver's
                return self._classify_source(node.func.value, seen)
        return "unknown"

    # ---- class-name-keyed registries (RFT) ------------------------------
    def visit_Attribute(self, node):
        if node.attr in ("__name__", "__qualname__"):
            self.add(node.lineno, "LOW", "rft",
                     f"uses .{node.attr}  -  if used as a registry key / dispatch, RFT "
                     f"rename mode changes it")
        self.generic_visit(node)

    def visit_ClassDef(self, node):
        for kw in node.keywords:
            if kw.arg == "metaclass":
                self.add(node.lineno, "LOW", "rft",
                         f"class '{node.name}' uses a metaclass  -  metaclass registries "
                         f"often key on class names; RFT rename mode may break them")
                break
        self.generic_visit(node)


def scan_file(path, opts):
    try:
        src = Path(path).read_text(encoding="utf-8", errors="replace")
    except OSError as e:
        return [(0, "HIGH", "unreadable", f"cannot read file: {e}")]
    try:
        tree = ast.parse(src, filename=str(path))
    except SyntaxError as e:
        return [(e.lineno or 0, "HIGH", "syntax-error",
                 f"does not parse ({e.msg})  -  PyArmor can't obfuscate it either")]
    # Pre-collect variable assignments (whole file) so eval/exec provenance can
    # trace a name back to its source. Imprecise (ignores scope/order) but
    # conservative: we only OMIT a finding when *every* assignment is local.
    assignments = {}
    for n in ast.walk(tree):
        if isinstance(n, ast.Assign):
            for t in n.targets:
                if isinstance(t, ast.Name):
                    assignments.setdefault(t.id, []).append(n.value)
        elif isinstance(n, ast.AnnAssign):
            if isinstance(n.target, ast.Name) and n.value is not None:
                assignments.setdefault(n.target.id, []).append(n.value)
        elif isinstance(n, ast.NamedExpr):       # walrus  (x := ...)
            if isinstance(n.target, ast.Name):
                assignments.setdefault(n.target.id, []).append(n.value)
        elif isinstance(n, (ast.With, ast.AsyncWith)):   # with open(...) as f
            for item in n.items:
                if isinstance(item.optional_vars, ast.Name):
                    assignments.setdefault(item.optional_vars.id, []).append(item.context_expr)
    sc = Scanner(path, opts, assignments)
    sc.visit(tree)
    # de-dup identical (line, category, message)
    seen = set()
    uniq = []
    for f in sc.findings:
        key = (f[0], f[2], f[3])
        if key not in seen:
            seen.add(key)
            uniq.append(f)
    return uniq


def _scan_worker(arg):
    """Top-level (picklable) worker for the process pool: (path, opts) -> (path, findings)."""
    path, opts = arg
    return str(path), scan_file(path, opts)


def write_xlsx(path, results, by_sev, by_cat, exclude_candidates, files_scanned):
    """Write findings + summary to an Excel .xlsx workbook. Requires openpyxl
    (raises ImportError if it isn't installed, so callers can degrade gracefully)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sev_fill = {
        "HIGH":   PatternFill("solid", fgColor="FFC7CE"),   # light red
        "MEDIUM": PatternFill("solid", fgColor="FFEB9C"),   # amber
        "LOW":    PatternFill("solid", fgColor="FFF2CC"),   # pale yellow
        "INFO":   PatternFill("solid", fgColor="E7E6E6"),   # gray
    }
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    wb = Workbook()

    # ── Findings sheet: one row per finding, worst-first within each file ──
    ws = wb.active
    ws.title = "Findings"
    headers = ["File", "Line", "Severity", "Category", "Message"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
    rows = [(p, line, sev, cat, msg)
            for p in sorted(results)
            for (line, sev, cat, msg) in results[p]]
    rows.sort(key=lambda r: (r[0], -SEVERITY_ORDER[r[2]], r[1]))
    for (p, line, sev, cat, msg) in rows:
        ws.append([p, line, sev, cat, msg])
        fill = sev_fill.get(sev)
        if fill:
            ws.cell(row=ws.max_row, column=3).fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
    for i, w in enumerate((50, 6, 9, 22, 90), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(row=1, column=5).alignment = Alignment(horizontal="left")

    # ── Summary sheet ─────────────────────────────────────────────────
    sm = wb.create_sheet("Summary")

    def _title(text):
        sm.append([text])
        sm.cell(row=sm.max_row, column=1).font = Font(bold=True)

    sm.append(["Files scanned", files_scanned])
    sm.append([])
    _title("Findings by severity")
    for k in ("HIGH", "MEDIUM", "LOW", "INFO"):
        sm.append([k, by_sev.get(k, 0)])
        fill = sev_fill.get(k)
        if fill:
            sm.cell(row=sm.max_row, column=1).fill = fill
    sm.append([])
    _title("Findings by category")
    for k, v in sorted(by_cat.items()):
        sm.append([k, v])
    sm.append([])
    _title("Modules to EXCLUDE from obfuscation (HIGH breakers)")
    for p in (exclude_candidates or ["(none)"]):
        sm.append([p])
    sm.column_dimensions["A"].width = 55
    sm.column_dimensions["B"].width = 12

    wb.save(path)


def iter_py_files(root, excludes):
    root = Path(root)
    if root.is_file():
        yield root
        return
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames
                       if d not in excludes and not d.startswith(".")
                       and d != "__pycache__"]
        for fn in filenames:
            if fn.endswith(".py"):
                p = Path(dirpath) / fn
                if not any(ex in p.parts for ex in excludes):
                    yield p


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Statically flag scripts PyArmor may break.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)
    ap.add_argument("root", help="file or directory to scan")
    ap.add_argument("--json", action="store_true", help="machine-readable output")
    ap.add_argument("-o", "--out", nargs="?", const="__AUTO__", default=None,
                    metavar="FILE",
                    help="also write the output to FILE (still prints to the terminal). "
                         "Bare -o auto-names it output.json (--json) or output.txt.")
    ap.add_argument("--xlsx", "--excel", dest="xlsx", nargs="?",
                    const="__AUTO__", default=None, metavar="FILE",
                    help="also write findings to an Excel .xlsx workbook (Findings + "
                         "Summary sheets), independent of --json/-o. Bare --xlsx/--excel "
                         "auto-names it output.xlsx. Requires openpyxl "
                         "(pip install openpyxl).")
    ap.add_argument("--min-severity", default="INFO",
                    choices=list(SEVERITY_ORDER), help="hide findings below this level")
    ap.add_argument("--fail-on", default="HIGH",
                    choices=list(SEVERITY_ORDER),
                    help="exit non-zero if any finding at/above this level (default HIGH)")
    ap.add_argument("--no-rft", action="store_true",
                    help="skip RFT/rename-mode checks (use if not using rename mode)")
    ap.add_argument("--no-bcc", action="store_true",
                    help="skip BCC/bytecode-to-C checks (use if not using BCC mode)")
    ap.add_argument("--no-trace-exec", action="store_true",
                    help="don't trace eval/exec input provenance; flag ALL exec/eval/compile "
                         "(default: omit those whose input traces to a literal/config/local file)")
    ap.add_argument("--checks", choices=["pyarmor", "all"], default="pyarmor",
                    help="'pyarmor' (default): only report failures OBFUSCATION INTRODUCES "
                         "(works plaintext, breaks obfuscated). 'all': also report pre-existing "
                         "business-logic/security issues (spawn/pool picklability, eval-of-input "
                         "RCE, sibling data-file loads) that fail the same with or without PyArmor.")
    ap.add_argument("--exclude", action="append", default=[],
                    help="directory name/part to skip (repeatable)")
    ap.add_argument("-j", "--jobs", type=int, default=0,
                    help="parallel worker processes (0 = auto = CPU count; 1 = serial)")
    opts = ap.parse_args(argv)

    min_sev = SEVERITY_ORDER[opts.min_severity]
    files = list(iter_py_files(opts.root, set(opts.exclude)))
    files_scanned = len(files)

    jobs = opts.jobs if opts.jobs > 0 else (os.cpu_count() or 1)
    # Process-pool overhead isn't worth it for a handful of files.
    if jobs > 1 and files_scanned > 16:
        from concurrent.futures import ProcessPoolExecutor
        chunk = max(1, files_scanned // (jobs * 4))
        with ProcessPoolExecutor(max_workers=jobs) as ex:
            scanned = ex.map(_scan_worker, ((p, opts) for p in files), chunksize=chunk)
    else:
        scanned = (_scan_worker((p, opts)) for p in files)

    results = {}   # path -> [findings]
    for path, raw in scanned:
        findings = [f for f in raw if SEVERITY_ORDER[f[1]] >= min_sev]
        if findings:
            findings.sort(key=lambda f: (-SEVERITY_ORDER[f[1]], f[0]))
            results[path] = findings

    # summary aggregates
    by_cat = {}
    by_sev = {k: 0 for k in SEVERITY_ORDER}
    exclude_candidates = set()
    for path, findings in results.items():
        for line, sev, cat, msg in findings:
            by_cat[cat] = by_cat.get(cat, 0) + 1
            by_sev[sev] += 1
            if sev == "HIGH" and cat in ("numba-jit", "func-serialization",
                                         "source-introspection", "trace-hooks"):
                exclude_candidates.add(path)

    worst = max((SEVERITY_ORDER[s] for f in results.values() for (_, s, _, _) in f),
                default=-1)
    exit_code = 1 if worst >= SEVERITY_ORDER[opts.fail_on] else 0

    # optional Excel workbook (independent of the text/JSON terminal output)
    if opts.xlsx is not None:
        xlsx_path = "output.xlsx" if opts.xlsx == "__AUTO__" else opts.xlsx
        try:
            write_xlsx(xlsx_path, results, by_sev, by_cat,
                       sorted(exclude_candidates), files_scanned)
            print(f"[written to {xlsx_path}]", file=sys.stderr)
        except ImportError:
            print("[--xlsx needs openpyxl: pip install openpyxl]", file=sys.stderr)
        except OSError as e:
            print(f"[could not write {xlsx_path}: {e}]", file=sys.stderr)

    # resolve the optional output file (bare -o auto-names by mode)
    if opts.out is None:
        outfile = None
    elif opts.out == "__AUTO__":
        outfile = "output.json" if opts.json else "output.txt"
    else:
        outfile = opts.out

    # build the whole output as one string, then print AND optionally save it
    if opts.json:
        out = {
            "files": {p: [{"line": l, "severity": s, "category": c, "message": m}
                          for (l, s, c, m) in fs] for p, fs in results.items()},
            "summary": {"by_severity": by_sev, "by_category": by_cat,
                        "exclude_candidates": sorted(exclude_candidates),
                        "files_scanned": files_scanned},
        }
        text = json.dumps(out, indent=2)
    elif not results:
        text = "No PyArmor-compatibility issues found."
    else:
        lines = []
        for path in sorted(results):
            lines.append(f"\n{path}")
            for line, sev, cat, msg in results[path]:
                lines.append(f"  {sev:6} [{cat}] line {line}: {msg}")
        lines.append("\n" + "=" * 70)
        lines.append("SUMMARY")
        lines.append("  findings by severity: " +
                     ", ".join(f"{k}={by_sev[k]}" for k in ["HIGH", "MEDIUM", "LOW", "INFO"]))
        lines.append("  findings by category: " +
                     ", ".join(f"{k}={v}" for k, v in sorted(by_cat.items())))
        if exclude_candidates:
            lines.append("\n  >> Modules to EXCLUDE from obfuscation (HIGH breakers):")
            for p in sorted(exclude_candidates):
                lines.append(f"       {p}")
        lines.append("\n  Static pre-filter only  -  confirm with encryptor/pyarmor_verify.py.")
        text = "\n".join(lines)

    print(text)
    if outfile:
        try:
            Path(outfile).write_text(text + "\n", encoding="utf-8")
            print(f"[written to {outfile}]", file=sys.stderr)
        except OSError as e:
            print(f"[could not write {outfile}: {e}]", file=sys.stderr)
    return exit_code


if __name__ == "__main__":
    sys.exit(main())
